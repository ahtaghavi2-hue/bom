from flask import Flask, render_template, request, jsonify, send_from_directory, send_file, redirect, flash, url_for
from flask_login import login_required, login_user, logout_user, current_user
import os
import json
import uuid
import re
from pathlib import Path
from io import BytesIO
from datetime import datetime

from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import PatternFill, Font, Alignment
import graphviz

from config import Config
from extensions import db, login_manager, mail, migrate, cors
from models import Product, Assembly, Part, Stage, Image, Document, Setting, User, ProductionSchedule, WorkOrder
from models import Manufacturer, ManufacturerEmail, ManufacturerPhone, ManufacturerSocial, StageDetail, part_manufacturers
from auth import auth_bp
from api_v2 import api_bp

app = Flask(__name__)
app.config.from_object(Config)

db.init_app(app)
login_manager.init_app(app)
mail.init_app(app)
migrate.init_app(app, db)
cors.init_app(app)

login_manager.login_view = 'auth.login'
login_manager.login_message = 'لطفاً ابتدا وارد سیستم شوید'

@login_manager.user_loader
def load_user(user_id):
    return db.session.get(User, int(user_id))

app.register_blueprint(auth_bp)
app.register_blueprint(api_bp)

UPLOAD_DIR = Path("uploads")
UPLOAD_DIR.mkdir(exist_ok=True)

# ───── DB ↔ Old-Format Compatibility Layer ─────

def _build_node_dict(obj, typ):
    children = []
    parent = None
    images = []
    stages = []
    notes = ''
    partCode = ''
    specs = ''
    docLink = ''
    partType = ''
    status = 'not_started'
    quantity = 0
    required_quantity = 1
    supplier = ''
    order_count = 0

    if typ == 'product':
        p = obj
        node_id = f'p{p.id}'
        parent = None
        notes = p.notes or ''
        partCode = p.code or ''
        specs = p.specs or ''
        order_count = p.order_count or 0
        children = [f'a{a.id}' for a in Assembly.query.filter_by(product_id=p.id, parent_id=None).all()]
        for a in Assembly.query.filter_by(product_id=p.id).all():
            for img in a.images.all():
                images.append({'url': img.url, 'label': img.label or 'طراحی'})
    elif typ == 'assembly':
        a = obj
        node_id = f'a{a.id}'
        parent = f'p{a.product_id}' if not a.parent_id else f'a{a.parent_id}'
        notes = a.description or ''
        children = [f'r{part.id}' for part in Part.query.filter_by(assembly_id=a.id).all()]
        for sub in Assembly.query.filter_by(parent_id=a.id).all():
            children.append(f'a{sub.id}')
        for img in a.images.all():
            images.append({'url': img.url, 'label': img.label or 'طراحی'})
    else:
        r = obj
        node_id = f'r{r.id}'
        parent = f'a{r.assembly_id}'
        notes = r.notes or ''
        partCode = r.part_code or ''
        specs = r.specs or ''
        partType = r.part_type or ''
        status = r.status or 'not_started'
        quantity = r.quantity or 0
        required_quantity = r.required_quantity or 1
        supplier = r.supplier or ''
        for s in r.stages.order_by(Stage.sort_order).all():
            stages.append({
                'name': s.name,
                'status': s.status or 'not_started',
                'estimated_material_cost': s.estimated_material_cost or 0,
                'actual_material_cost': s.actual_material_cost or 0,
                'estimated_labor_cost': s.estimated_labor_cost or 0,
                'actual_labor_cost': s.actual_labor_cost or 0,
                'estimated_overhead': s.estimated_overhead or 0,
                'actual_overhead': s.actual_overhead or 0,
                'estimated_hours': s.estimated_hours or 0,
                'actual_hours': s.actual_hours or 0,
            })
        for img in r.images.all():
            images.append({'url': img.url, 'label': img.label or 'طراحی'})

    return {
        'id': node_id, 'name': obj.name, 'type': typ,
        'parent': parent, 'children': children,
        'notes': notes, 'images': images,
        'partCode': partCode, 'specs': specs, 'docLink': docLink,
        'partType': partType, 'status': status,
        'quantity': quantity, 'required_quantity': required_quantity,
        'stages': stages, 'supplier': supplier, 'supplier_email': r.supplier_email or '' if typ == 'part' else '',
        'order_count': order_count
    }

def load_data():
    nodes = {}
    root_ids = []
    for p in Product.query.all():
        nd = _build_node_dict(p, 'product')
        nodes[nd['id']] = nd
        root_ids.append(nd['id'])
    for a in Assembly.query.all():
        nd = _build_node_dict(a, 'assembly')
        nodes[nd['id']] = nd
    for r in Part.query.all():
        nd = _build_node_dict(r, 'part')
        nodes[nd['id']] = nd
    return {'nodes': nodes, 'root_ids': root_ids}

def _resolve_node(node_id):
    if node_id.startswith('p'):
        p = db.session.get(Product, int(node_id[1:]))
        return p, 'product' if p else None
    elif node_id.startswith('a'):
        a = db.session.get(Assembly, int(node_id[1:]))
        return a, 'assembly' if a else None
    elif node_id.startswith('r'):
        r = db.session.get(Part, int(node_id[1:]))
        return r, 'part' if r else None
    return None, None

def _parent_node_id(obj, typ):
    if typ == 'assembly':
        a = obj
        if a.parent_id:
            return f'a{a.parent_id}'
        return f'p{a.product_id}'
    elif typ == 'part':
        return f'a{obj.assembly_id}'
    return None

def get_ancestor_order_count(node_id, data):
    node = data['nodes'].get(node_id)
    if not node:
        return 1
    while node:
        if node.get('type') == 'product' and node.get('order_count', 0) > 0:
            return node['order_count']
        if node.get('parent'):
            node = data['nodes'].get(node['parent'])
        else:
            break
    return 1

# ───── Classic Auth Routes ─────

@app.route('/login', methods=['GET', 'POST'])
def login_page():
    if request.method == 'GET':
        return render_template('modern/login.html')
    data = request.get_json(silent=True) or request.form
    username = data.get('username', '').strip()
    password = data.get('password', '')
    user = User.query.filter((User.username == username) | (User.email == username)).first()
    if not user or not user.check_password(password):
        return render_template('modern/login.html', error='نام کاربری یا رمز عبور اشتباه است')
    login_user(user)
    user.last_login = datetime.utcnow()
    db.session.commit()
    return redirect(url_for('index'))

@app.route('/logout')
def logout_page():
    logout_user()
    return redirect(url_for('login_page'))

# ───── Classic Routes ─────

@app.route('/')
@login_required
def index():
    return render_template('index.html')

@app.route('/api/data', methods=['GET'])
@login_required
def get_data():
    return jsonify(load_data())

@app.route('/api/notifications')
@login_required
def get_notifications():
    data = load_data()
    notifications = {'shortage': [], 'orders_pending': [], 'in_progress': []}
    for node_id, node in data['nodes'].items():
        if node['type'] == 'part':
            order_count = get_ancestor_order_count(node_id, data)
            total_req = node.get('required_quantity', 1) * order_count
            available = node.get('quantity', 0)
            if available < total_req:
                notifications['shortage'].append({
                    'id': node_id, 'name': node['name'],
                    'required': total_req, 'available': available,
                    'shortage': total_req - available
                })
            status = node.get('status', 'not_started')
            if status == 'in_progress':
                notifications['in_progress'].append({
                    'id': node_id, 'name': node['name'],
                    'stages': len(node.get('stages', []))
                })
        if node['type'] == 'product' and node.get('order_count', 0) > 0:
            notifications['orders_pending'].append({
                'id': node_id, 'name': node['name'], 'count': node['order_count']
            })
    return jsonify(notifications)

@app.route('/api/node', methods=['POST'])
@login_required
def create_node():
    req = request.json
    typ = req.get('type', 'assembly')
    parent_id = req.get('parent')
    name = req.get('name', 'بدون نام')

    if typ == 'product':
        p = Product(name=name, code='', specs='', notes='', order_count=0, status='active')
        db.session.add(p)
        db.session.flush()
    elif typ == 'assembly':
        parent_obj, parent_typ = _resolve_node(parent_id)
        if parent_typ == 'product':
            a = Assembly(product_id=parent_obj.id, parent_id=None, name=name)
        else:
            a = Assembly(product_id=parent_obj.product_id if hasattr(parent_obj, 'product_id') else 1, parent_id=parent_obj.id, name=name)
        db.session.add(a)
        db.session.flush()
    else:
        parent_obj, parent_typ = _resolve_node(parent_id)
        assembly_id = parent_obj.id if parent_typ == 'assembly' else 1
        r = Part(assembly_id=assembly_id, name=name, part_code='', specs='', part_type='make',
                 quantity=0, required_quantity=1, supplier='', notes='', status='not_started')
        db.session.add(r)
        db.session.flush()

    db.session.commit()
    data = load_data()
    nd = data['nodes'].get(f'{typ[0]}{p.id if typ=="product" else a.id if typ=="assembly" else r.id}')
    return jsonify({'success': True, 'node': nd})

@app.route('/api/node/<node_id>', methods=['PUT'])
@login_required
def update_node(node_id):
    req = request.json
    obj, typ = _resolve_node(node_id)
    if not obj:
        return jsonify({'error': 'Not found'}), 404

    if typ == 'product':
        for key in ['name', 'notes', 'order_count']:
            if key in req:
                setattr(obj, key if key != 'order_count' else 'order_count', req[key])
        obj.code = req.get('partCode', obj.code)
        obj.specs = req.get('specs', obj.specs)
    elif typ == 'assembly':
        if 'name' in req: obj.name = req['name']
        obj.description = req.get('notes', obj.description)
    else:
        for key in ['name', 'notes', 'partCode', 'specs', 'partType', 'status', 'supplier']:
            if key in req:
                setattr(obj, key, req[key])
        if 'quantity' in req: obj.quantity = int(req['quantity'])
        if 'required_quantity' in req: obj.required_quantity = int(req['required_quantity'])
        if 'stages' in req:
            Stage.query.filter_by(part_id=obj.id).delete()
            for idx, s in enumerate(req['stages']):
                st = Stage(part_id=obj.id, name=s.get('name', 'Stage'),
                           status=s.get('status', 'not_started'), sort_order=idx,
                           estimated_material_cost=s.get('estimated_material_cost', 0),
                           actual_material_cost=s.get('actual_material_cost', 0),
                           estimated_labor_cost=s.get('estimated_labor_cost', 0),
                           actual_labor_cost=s.get('actual_labor_cost', 0),
                           estimated_overhead=s.get('estimated_overhead', 0),
                           actual_overhead=s.get('actual_overhead', 0),
                           estimated_hours=s.get('estimated_hours', 0),
                           actual_hours=s.get('actual_hours', 0))
                db.session.add(st)
        if 'images' in req:
            Image.query.filter_by(part_id=obj.id).delete()
            for img in req['images']:
                im = Image(part_id=obj.id, url=img.get('url', ''), label=img.get('label', ''))
                db.session.add(im)

    db.session.commit()
    return jsonify({'success': True, 'node': _build_node_dict(obj, typ)})

@app.route('/api/node/<node_id>', methods=['PATCH'])
@login_required
def quick_update_node(node_id):
    req = request.json
    obj, typ = _resolve_node(node_id)
    if not obj:
        return jsonify({'error': 'Not found'}), 404
    if 'name' in req:
        old_name = obj.name
        obj.name = req['name']
    if 'quantity' in req and typ == 'part':
        obj.quantity = int(req['quantity'])
    if 'required_quantity' in req and typ == 'part':
        obj.required_quantity = int(req['required_quantity'])
    db.session.commit()
    return jsonify({'success': True, 'node': _build_node_dict(obj, typ)})

@app.route('/api/node/<node_id>/move', methods=['POST'])
@login_required
def move_node(node_id):
    obj, typ = _resolve_node(node_id)
    if not obj:
        return jsonify({'error': 'Not found'}), 404
    req = request.json
    new_parent_id = req.get('new_parent')
    new_parent_obj, new_parent_typ = _resolve_node(new_parent_id) if new_parent_id and new_parent_id != '#' else (None, None)

    if typ == 'assembly':
        old_parent_id = obj.parent_id
        if new_parent_typ == 'product':
            obj.parent_id = None
            obj.product_id = new_parent_obj.id
        elif new_parent_typ == 'assembly':
            obj.parent_id = new_parent_obj.id
            obj.product_id = new_parent_obj.product_id
        else:
            obj.parent_id = None
    elif typ == 'part':
        if new_parent_typ == 'assembly':
            old_assembly_id = obj.assembly_id
            obj.assembly_id = new_parent_obj.id

    db.session.commit()
    return jsonify({'success': True})

@app.route('/api/node/<node_id>', methods=['DELETE'])
@login_required
def delete_node(node_id):
    obj, typ = _resolve_node(node_id)
    if not obj:
        return jsonify({'error': 'Not found'}), 404

    def delete_recursive(o, t):
        if t == 'product':
            for a in Assembly.query.filter_by(product_id=o.id).all():
                delete_recursive(a, 'assembly')
            db.session.delete(o)
        elif t == 'assembly':
            for sub in Assembly.query.filter_by(parent_id=o.id).all():
                delete_recursive(sub, 'assembly')
            for p in Part.query.filter_by(assembly_id=o.id).all():
                delete_recursive(p, 'part')
            db.session.delete(o)
        elif t == 'part':
            Stage.query.filter_by(part_id=o.id).delete()
            Image.query.filter_by(part_id=o.id).delete()
            db.session.delete(o)

    delete_recursive(obj, typ)
    db.session.commit()
    return jsonify({'success': True})

@app.route('/api/upload', methods=['POST'])
@login_required
def upload():
    if 'file' not in request.files:
        return jsonify({'error': 'No file'}), 400
    file = request.files['file']
    if file.filename == '':
        return jsonify({'error': 'No file'}), 400
    ext = Path(file.filename).suffix
    filename = f"{uuid.uuid4().hex[:8]}{ext}"
    filepath = UPLOAD_DIR / filename
    file.save(filepath)
    return jsonify({'url': f'/uploads/{filename}', 'filename': filename})

@app.route('/uploads/<path:filename>')
def uploaded_file(filename):
    return send_from_directory(UPLOAD_DIR, filename)

@app.route('/api/export/excel')
@login_required
def export_excel():
    product_id = request.args.get('product_id')
    data = load_data()
    if not product_id or product_id not in data['nodes']:
        return "محصول انتخاب نشده است", 400
    product = data['nodes'][product_id]
    target_qty = max(product.get('order_count', 1), 1)
    def get_real_status(node):
        if node.get('type') == 'part':
            required = node.get('required_quantity', 1)
            total_req = required * target_qty
            available = node.get('quantity', 0)
            if available >= total_req: return 'کافی'
            elif available > 0: return 'ناقص'
            else: return 'کسری'
        return 'ندارد'
    wb = Workbook()
    wb.remove(wb.active)
    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    def add_sheet_for_node(node_id, sheet_name, parent_total_req=1):
        node = data['nodes'].get(node_id)
        if not node: return
        clean_name = re.sub(r'[\\\/\?\*\[\]]', '-', sheet_name)[:31]
        ws = wb.create_sheet(title=clean_name)
        headers = ['نام قطعه', 'کد فنی', 'تصویر', 'مراحل ساخت', 'نیاز کل (سفارش)', 'موجودی', 'وضعیت', 'تأمین‌کننده', 'ایمیل تأمین‌کننده', 'هزینه مواد', 'دستمزد', 'سربار', 'مجموع هزینه', 'توضیحات']
        ws.append(headers)
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="4CAF50", end_color="4CAF50", fill_type="solid")
            cell.alignment = Alignment(horizontal='center')
        row_idx = 2
        def write_nodes_recursive(nid, level, parent_req):
            nonlocal row_idx
            child_node = data['nodes'].get(nid)
            if not child_node: return
            total_required = child_node.get('required_quantity', 1) * parent_req
            stages_str = " - ".join([s['name'] for s in child_node.get('stages', [])]) if child_node.get('stages') else "-"
            indent = "  " * level
            ws.cell(row=row_idx, column=1, value=indent + child_node['name'])
            ws.cell(row=row_idx, column=2, value=child_node.get('partCode', '-'))
            if child_node.get('images') and len(child_node['images']) > 0:
                img_filename = child_node['images'][0]['url'].split('/')[-1]
                img_path = os.path.abspath(UPLOAD_DIR / img_filename)
                if os.path.exists(img_path):
                    try:
                        img = XLImage(img_path)
                        img.width = 40; img.height = 40
                        ws.add_image(img, f"C{row_idx}")
                    except Exception: pass
            ws.cell(row=row_idx, column=4, value=stages_str)
            ws.cell(row=row_idx, column=5, value=total_required)
            qty = child_node.get('quantity', 0)
            cell_qty = ws.cell(row=row_idx, column=6, value=qty)
            if qty < total_required: cell_qty.fill = red_fill
            elif qty >= total_required: cell_qty.fill = green_fill
            ws.cell(row=row_idx, column=7, value=get_real_status(child_node))
            ws.cell(row=row_idx, column=8, value=child_node.get('supplier', ''))
            ws.cell(row=row_idx, column=9, value=child_node.get('supplier_email', ''))
            stages_cost = child_node.get('stages', [])
            total_mat = sum(s.get('estimated_material_cost', 0) for s in stages_cost)
            total_labor = sum(s.get('estimated_labor_cost', 0) for s in stages_cost)
            total_overhead = sum(s.get('estimated_overhead', 0) for s in stages_cost)
            ws.cell(row=row_idx, column=10, value=total_mat)
            ws.cell(row=row_idx, column=11, value=total_labor)
            ws.cell(row=row_idx, column=12, value=total_overhead)
            ws.cell(row=row_idx, column=13, value=total_mat + total_labor + total_overhead)
            ws.cell(row=row_idx, column=14, value=child_node.get('notes', ''))
            ws.row_dimensions[row_idx].height = 45
            row_idx += 1
            for grandchild_id in child_node.get('children', []):
                write_nodes_recursive(grandchild_id, level + 1, total_required)
        for child_id in node.get('children', []):
            write_nodes_recursive(child_id, 0, parent_total_req)
        for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M']:
            ws.column_dimensions[col].width = 18
        ws.column_dimensions['N'].width = 35
    if product_id and product_id in data['nodes']:
        selected_node = data['nodes'][product_id]
        add_sheet_for_node(product_id, selected_node['name'], target_qty)
        def add_child_sheets(node_id):
            node = data['nodes'].get(node_id)
            if not node: return
            for child_id in node.get('children', []):
                child = data['nodes'].get(child_id)
                if child and child['type'] == 'assembly':
                    add_sheet_for_node(child_id, child['name'], target_qty)
                add_child_sheets(child_id)
        add_child_sheets(product_id)
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    filename = f'BOM_{product_id}_OrderQty{target_qty}.xlsx'
    return send_file(output, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', as_attachment=True, download_name=filename)

@app.route('/api/export/schematic')
@login_required
def export_schematic():
    product_id = request.args.get('product_id')
    data = load_data()
    if not product_id or product_id not in data['nodes']:
        return "محصول انتخاب نشده است", 400
    root_node = data['nodes'][product_id]
    root_order_count = max(root_node.get('order_count', 1), 1)
    def get_real_status(node):
        if node.get('type') == 'part':
            required = node.get('required_quantity', 1)
            total_req = required * root_order_count
            available = node.get('quantity', 0)
            if available >= total_req: return 'completed'
            elif available > 0: return 'in_progress'
            else: return 'not_started'
        return node.get('status', 'not_started')
    dot = graphviz.Digraph('ProductSchematic', format='pdf')
    dot.attr(rankdir='LR', size='30,30!', dpi='300', bgcolor='#FFFFFF')
    dot.attr('node', shape='plaintext', fontname='Tahoma')
    dot.attr('edge', penwidth='3')
    def get_edge_color(node):
        status = get_real_status(node)
        if status == 'completed': return '#4CAF50'
        elif status == 'in_progress': return '#FF9800'
        else: return '#F44336'
    def get_stage_color(stage):
        status = stage.get('status', 'not_started')
        if status == 'completed': return '#4CAF50'
        elif status == 'in_progress': return '#FF9800'
        else: return '#F44336'
    def create_node_label(node):
        img_html = ""
        if node.get('images') and len(node['images']) > 0:
            img_filename = node['images'][0]['url'].split('/')[-1]
            img_path = os.path.abspath(UPLOAD_DIR / img_filename)
            if os.path.exists(img_path):
                img_html = f'<TR><TD><IMG SRC="{img_path}" SCALE="TRUE" WIDTH="120"/></TD></TR>'
        stages_html = ""
        if node.get('stages'):
            stages_html = '<TR><TD><TABLE BORDER="0" CELLBORDER="1" CELLSPACING="0" CELLPADDING="4"><TR>'
            for s in node['stages']:
                color = get_stage_color(s)
                stages_html += f'<TD BGCOLOR="{color}"><FONT COLOR="#FFFFFF" POINT-SIZE="10">{s["name"]}</FONT></TD>'
            stages_html += '</TR></TABLE></TD></TR>'
        total_req_display = (node.get('required_quantity', 1) * root_order_count) if node.get('type') == 'part' else node.get('required_quantity', 1)
        qty_info = f"موجودی: {node.get('quantity', 0)} | نیاز کل: {total_req_display}"
        stages_cost = node.get('stages', [])
        total_cost = sum(s.get('estimated_material_cost', 0) + s.get('estimated_labor_cost', 0) + s.get('estimated_overhead', 0) for s in stages_cost)
        cost_html = f'<TR><TD><FONT POINT-SIZE="10" COLOR="#2C3E50">هزینه برآوردی: {total_cost:,}</FONT></TD></TR>' if total_cost else ''
        supplier_html = f'<TR><TD><FONT POINT-SIZE="10" COLOR="#555555">تأمین‌کننده: {node.get("supplier", "-")}</FONT></TD></TR>' if node.get('supplier') else ''
        label = f'''<<TABLE BORDER="1" CELLBORDER="0" CELLSPACING="4" BGCOLOR="#F8F9FA" CELLPADDING="8">
            {img_html}
            <TR><TD><B><FONT POINT-SIZE="14" COLOR="#2C3E50">{node['name']}</FONT></B></TD></TR>
            <TR><TD><FONT POINT-SIZE="11" COLOR="#555555">کد فنی: {node.get('partCode', '-')}</FONT></TD></TR>
            <TR><TD><FONT POINT-SIZE="11" COLOR="#555555">{qty_info}</FONT></TD></TR>
            {cost_html}
            {supplier_html}
            {stages_html}
        </TABLE>>'''
        return label
    def add_nodes_to_graph(nid, parent_id=None):
        n = data['nodes'][nid]
        label = create_node_label(n)
        safe_id = f"node_{nid}"
        dot.node(safe_id, label=label)
        if parent_id:
            edge_color = get_edge_color(n)
            dot.edge(f"node_{parent_id}", safe_id, color=edge_color)
        for child_id in n.get('children', []):
            add_nodes_to_graph(child_id, nid)
    add_nodes_to_graph(product_id)
    pdf_output = dot.pipe()
    return send_file(BytesIO(pdf_output), mimetype='application/pdf',
                     as_attachment=True, download_name=f'Schematic_{root_node["name"]}.pdf')

# ───── Scheduling Endpoints ─────

@app.route('/api/schedules')
@login_required
def list_schedules():
    schedules = ProductionSchedule.query.order_by(ProductionSchedule.created_at.desc()).all()
    return jsonify({'success': True, 'data': [{
        'id': s.id, 'product_id': s.product_id, 'product_name': s.product.name if s.product else '',
        'quantity': s.quantity, 'status': s.status, 'priority': s.priority,
        'start_date': s.start_date.isoformat() if s.start_date else None,
        'end_date': s.end_date.isoformat() if s.end_date else None,
        'created_at': s.created_at.isoformat(), 'notes': s.notes
    } for s in schedules]})

@app.route('/api/schedules', methods=['POST'])
@login_required
def create_schedule():
    data = request.json
    s = ProductionSchedule(
        product_id=data['product_id'],
        quantity=data.get('quantity', 1),
        status=data.get('status', 'planned'),
        priority=data.get('priority', 'normal'),
        start_date=datetime.fromisoformat(data['start_date']) if data.get('start_date') else None,
        end_date=datetime.fromisoformat(data['end_date']) if data.get('end_date') else None,
        notes=data.get('notes', ''),
        created_by=current_user.id
    )
    db.session.add(s)
    db.session.commit()
    _auto_work_orders(s)
    return jsonify({'success': True, 'id': s.id})

def _auto_work_orders(schedule):
    for part in Part.query.filter(Part.assembly_id.in_(
        db.session.query(Assembly.id).filter(Assembly.product_id == schedule.product_id)
    )).all():
        wo = WorkOrder(schedule_id=schedule.id, part_id=part.id,
                       quantity=part.required_quantity * schedule.quantity,
                       status='pending', due_date=schedule.end_date)
        db.session.add(wo)
    db.session.commit()

# ───── Settings ─────

@app.route('/settings', methods=['GET', 'POST'])
@login_required
def settings_page():
    if request.method == 'POST':
        for key in ['mail_server', 'mail_port', 'mail_username', 'mail_password', 'mail_use_tls']:
            val = request.form.get(key, '')
            if key == 'mail_port':
                try: val = str(int(val))
                except: val = '587'
            Setting.set(key, val)
        flash('تنظیمات با موفقیت ذخیره شد', 'success')
        return redirect(url_for('settings_page'))
    settings = {
        'mail_server': Setting.get('mail_server', 'smtp.gmail.com'),
        'mail_port': Setting.get('mail_port', '587'),
        'mail_username': Setting.get('mail_username', ''),
        'mail_password': Setting.get('mail_password', ''),
        'mail_use_tls': Setting.get('mail_use_tls', 'true'),
    }
    return render_template('settings.html', settings=settings)

@app.route('/api/test-email', methods=['POST'])
@login_required
def test_email():
    data = request.json
    if current_user.role not in ('admin', 'engineer'):
        return jsonify({'message': 'دسترسی غیرمجاز'}), 403
    try:
        from flask_mail import Message
        mail_server = data.get('mail_server', Setting.get('mail_server', ''))
        mail_port = data.get('mail_port', int(Setting.get('mail_port', '587')))
        mail_username = data.get('mail_username', Setting.get('mail_username', ''))
        mail_password = data.get('mail_password', Setting.get('mail_password', ''))
        mail_use_tls = data.get('mail_use_tls', Setting.get('mail_use_tls', 'true')) in (True, 'true')
        if not mail_server or not mail_username:
            return jsonify({'message': 'لطفاً ابتدا سرور SMTP و نام کاربری را تنظیم کنید'}), 400
        app.config['MAIL_SERVER'] = mail_server
        app.config['MAIL_PORT'] = mail_port
        app.config['MAIL_USE_TLS'] = mail_use_tls
        app.config['MAIL_USERNAME'] = mail_username
        app.config['MAIL_PASSWORD'] = mail_password
        app.config['MAIL_DEFAULT_SENDER'] = mail_username
        msg = Message(
            subject='تست ارسال ایمیل - BOM Manager',
            recipients=[mail_username],
            body='این یک ایمیل آزمایشی از سامانه مدیریت BOM است.\n\nاگر این ایمیل را دریافت کرده‌اید، تنظیمات SMTP شما به درستی کار می‌کند.'
        )
        mail.send(msg)
        return jsonify({'message': 'ایمیل آزمایشی با موفقیت ارسال شد. صندوق ورودی خود را بررسی کنید.'})
    except Exception as e:
        return jsonify({'message': f'خطا در ارسال: {str(e)}'}), 500

@app.route('/api/send-part-email', methods=['POST'])
@login_required
def send_part_email():
    data = request.json
    part_id = data.get('part_id')
    obj, typ = _resolve_node(part_id)
    if not obj or typ != 'part':
        return jsonify({'message': 'قطعه یافت نشد'}), 404
    email_to = data.get('email', obj.supplier_email)
    if not email_to:
        return jsonify({'message': 'ایمیل تأمین‌کننده تنظیم نشده است'}), 400
    try:
        from email_notif import send_low_stock_email
        success = send_low_stock_email(obj)
        if success:
            return jsonify({'message': f'ایمیل هشدار کمبود به {email_to} ارسال شد'})
        else:
            return jsonify({'message': 'ارسال ایمیل ناموفق بود. تنظیمات SMTP را بررسی کنید.'}), 500
    except Exception as e:
        return jsonify({'message': f'خطا: {str(e)}'}), 500

# ───── Documents ─────

@app.route('/api/documents/<node_id>', methods=['GET'])
@login_required
def list_documents(node_id):
    obj, typ = _resolve_node(node_id)
    if not obj:
        return jsonify({'success': False}), 404
    docs = Document.query.filter_by(part_id=obj.id if typ=='part' else None,
                                     assembly_id=obj.id if typ=='assembly' else None).all()
    return jsonify({'success': True, 'data': [{
        'id': d.id, 'filename': d.original_name or d.filename,
        'file_type': d.file_type, 'file_size': d.file_size,
        'url': d.url, 'uploaded_at': d.uploaded_at.isoformat(),
        'notes': d.notes
    } for d in docs]})

@app.route('/api/documents/upload/<node_id>', methods=['POST'])
@login_required
def upload_document(node_id):
    obj, typ = _resolve_node(node_id)
    if not obj:
        return jsonify({'error': 'Node not found'}), 404
    if 'file' not in request.files:
        return jsonify({'error': 'No file'}), 400
    file = request.files['file']
    if file.filename == '':
        return jsonify({'error': 'No file'}), 400
    ext = Path(file.filename).suffix.lower()
    original_name = file.filename
    filename = f"doc_{uuid.uuid4().hex[:8]}{ext}"
    filepath = UPLOAD_DIR / filename
    file.save(filepath)
    file_size = os.path.getsize(filepath)
    type_map = {
        '.pdf': 'pdf', '.step': 'cad', '.stp': 'cad',
        '.dwg': 'cad', '.dxf': 'cad', '.stl': 'cad',
        '.sldprt': 'cad', '.sldasm': 'cad', '.iges': 'cad',
        '.igs': 'cad', '.3dm': 'cad', '.obj': 'cad',
        '.zip': 'archive', '.doc': 'doc', '.docx': 'doc',
        '.xls': 'xls', '.xlsx': 'xls', '.jpg': 'image',
        '.jpeg': 'image', '.png': 'image'
    }
    doc = Document(
        part_id=obj.id if typ == 'part' else None,
        assembly_id=obj.id if typ == 'assembly' else None,
        filename=filename, original_name=original_name,
        file_type=type_map.get(ext, 'other'), file_size=file_size,
        url=f'/uploads/{filename}'
    )
    db.session.add(doc)
    db.session.commit()
    return jsonify({'success': True, 'document': {
        'id': doc.id, 'filename': original_name,
        'file_type': doc.file_type, 'file_size': file_size,
        'url': doc.url, 'uploaded_at': doc.uploaded_at.isoformat()
    }})

@app.route('/api/documents/<int:doc_id>', methods=['DELETE'])
@login_required
def delete_document(doc_id):
    doc = db.session.get(Document, doc_id)
    if not doc:
        return jsonify({'error': 'Not found'}), 404
    filepath = UPLOAD_DIR / doc.filename
    if filepath.exists():
        filepath.unlink()
    db.session.delete(doc)
    db.session.commit()
    return jsonify({'success': True})

# ───── Init ─────

def init_db():
    with app.app_context():
        db.create_all()
        if not User.query.filter_by(username='admin').first():
            admin = User(username='admin', email='admin@bom-system.com', role='admin')
            admin.set_password('admin123')
            db.session.add(admin)
            db.session.commit()
            print('[OK] Default admin user created (admin/admin123)')

if __name__ == '__main__':
    init_db()
    from apscheduler.schedulers.background import BackgroundScheduler
    scheduler = BackgroundScheduler()
    scheduler.add_job(
        func=lambda: __import__('email_notif', fromlist=['check_and_notify_low_stock']).check_and_notify_low_stock(),
        trigger='interval', hours=6, id='low_stock_check'
    )
    scheduler.start()
    app.run(debug=True, port=5000)
